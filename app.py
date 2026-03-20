import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import io

def create_timetable():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # --- Styles ---
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    teal_fill = PatternFill(start_color="78C5A8", fill_type="solid")
    grey_fill = PatternFill(start_color="EBEBEB", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B183", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # --- Setup Columns ---
    col_widths = {'A': 12, 'B': 16, 'C': 16, 'D': 14, 'E': 14, 'F': 16, 'G': 14, 'H': 14, 'I': 16, 'J': 16}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    def style_range(cell_range, fill=None, font=None):
        for row in ws[cell_range]:
            for cell in row:
                cell.alignment = center_align
                cell.border = border
                if fill: cell.fill = fill
                if font: cell.font = font

    # --- WEEK 1 ---
    ws.append(["Week 1", "Form", "Period 1", "Break/Period 2", "", "Period 3", "Period 4/Lunch", "", "Lesson 5", "Form"])
    ws.append(["", "8.30-8.50", "8.50-9.50", "9.50-11.05", "", "11.05-12.05", "12.05-13.50", "", "13.50-14.50", "14.50-15.10"])
    ws.merge_cells('D1:E1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('D2:E2')
    ws.merge_cells('G2:H2')
    style_range('A1:J2', font=bold_font)

    ws.append(["Monday", "11SIN2 E1.15 Miss Murphy", "", "9.50-10.05\nBreak", "10.05-11.05 10Y2\nE2.18 Miss Murphy", "9Y2 E2.9 Mr\nMupfumbati", "12.05-13.05 8X2\nE2.18 Miss Murphy", "13.05-13.50\nLunch", "", "11SIN2 E1.15 Miss Murphy"])
    ws.append(["Tuesday", "11SIN2 E1.15 Miss Murphy", "8X2 E2.18 Miss\nMurphy", "9.50-10.50 9Y2\nE2.9 Mr Mupfumbati", "10.50-11.05\nBreak", "9X2 E2.15 Mrs\nBouf-Tah", "", "13.05-13.50\nLunch", "10Y2 E2.18 Miss\nMurphy", "11SIN2 E1.15 Miss Murphy"])
    ws.merge_cells('G4:G5') 
    ws.append(["Wednesday", "11SIN2 E1.15 Miss Murphy", "10Y2 E2.18 Miss\nMurphy", "9.50-10.50 7Y3\nE2.15 Mrs Bouf-Tah", "10.50-11.05\nBreak", "9Y2 E2.9 Mr\nMupfumbati", "", "13.05-13.50\nLunch", "8X2 E2.18 Miss\nMurphy", "11SIN2 E1.15 Miss Murphy"])
    ws.merge_cells('G5:G6')
    ws.append(["Thursday", "11SIN2 E1.15 Miss Murphy", "", "9.50-10.05\nBreak", "Mentor Meeting\nE2.18", "", "12.05-12.50\nLunch", "12.50-13.50 10Y2\nE2.18 Miss Murphy", "", "11SIN2 E1.15 Miss Murphy"])
    ws.append(["Friday", "University"])
    ws.merge_cells('B7:J7')

    # --- WEEK 2 ---
    ws.append([]) 
    ws.append(["Week 2", "Form", "Period 1", "Break/Period 2", "", "Period 3", "Period 4/Lunch", "", "Lesson 5", "Form"])
    ws.append(["", "8.30-8.50", "8.50-9.50", "9.50-11.05", "", "11.05-12.05", "12.05-13.50", "", "13.50-14.50", "14.50-15.10"])
    ws.merge_cells('D9:E9')
    ws.merge_cells('G9:H9')
    ws.merge_cells('D10:E10')
    ws.merge_cells('G10:H10')
    style_range('A9:J10', font=bold_font)

    ws.append(["Monday", "11SIN2 E1.15 Miss Murphy", "", "9.50-10.05\nBreak", "", "9Y2 E2.9 Mr\nMupfumbati", "12.05-13.05 8X2\nE2.18 Miss Murphy", "13.05-13.50\nLunch", "10Y2 E2.18 Miss\nMurphy", "11SIN2 E1.15 Miss Murphy"])
    ws.append(["Tuesday", "11SIN2 E1.15 Miss Murphy", "7Y3 E2.15 Mrs\nBouf-Tah", "9.50-10.50 9Y2\nE2.9 Mr Mupfumbati", "10.50-11.05\nBreak", "", "12.05-13.05 8X2\nE2.18 Miss Murphy", "13.05-13.50\nLunch", "10Y2 E2.18 Miss\nMurphy", "11SIN2 E1.15 Miss Murphy"])
    ws.append(["Wednesday", "11SIN2 E1.15 Miss Murphy", "10Y2 E2.18 Miss\nMurphy", "9.50-10.05\nBreak", "", "", "", "13.05-13.50\nLunch", "9X2 E2.15 Mrs\nBouf-Tah", "11SIN2 E1.15 Miss Murphy"])
    ws.append(["Thursday", "11SIN2 E1.15 Miss Murphy", "9Y2 E2.9 Mr\nMupfumbati", "9.50-10.05\nBreak", "", "Mentor Meeting\nE2.18", "12.05-13.05 8X2\nE2.18 Miss Murphy", "13.05-13.50\nLunch", "7Y3 E2.15 Mrs\nBouf-Tah", "11SIN2 E1.15 Miss Murphy"])
    ws.append(["Friday", "University"])
    ws.merge_cells('B15:J15')

    # --- Apply Global Styles and Colors ---
    style_range('A1:J15') 
    for row in range(3, 7):
        ws[f'B{row}'].fill = teal_fill
        ws[f'J{row}'].fill = teal_fill
    for row in range(11, 15):
        ws[f'B{row}'].fill = teal_fill
        ws[f'J{row}'].fill = teal_fill

    style_range('A7:J7', fill=orange_fill, font=bold_font)
    style_range('A15:J15', fill=orange_fill, font=bold_font)

    grey_cells = ['D3', 'E4', 'E5', 'D6', 'G6', 'H3', 'H4', 'H5', 'H6', 'D11', 'E12', 'D13', 'D14', 'H11', 'H12', 'H13', 'H14']
    for cell in grey_cells:
        ws[cell].fill = grey_fill
        
    for row in range(1, 16):
        ws.row_dimensions[row].height = 40 if row not in [1, 2, 7, 8, 9, 10, 15] else 20

    # Save to memory instead of hard drive
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- Streamlit Web App Interface ---
st.set_page_config(page_title="Timetable Generator", page_icon="🏫")

st.title("🏫 School Timetable Generator")
st.write("Click the button below to generate and download your formatted Excel timetable.")

# Create the download button
excel_file = create_timetable()

st.download_button(
    label="📥 Download Timetable (Excel)",
    data=excel_file,
    file_name="School_Timetable.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
